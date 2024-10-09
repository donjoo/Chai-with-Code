import os
import requests
from bs4 import BeautifulSoup  # type: ignore
from openpyxl import load_workbook # type: ignore


def extract_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    title = soup.find('title').get_text()
    paragraphs=soup.find_all('p')
    article_text = ' '.join([p.get_text() for p in paragraphs])

    return title, article_text



# for creating unique dir ,while running program each time.
def create_unique_directory(base_name):
    if not os.path.exists(base_name):
        os.makedirs(base_name)

        return base_name

    index = 1
    while True:
        new_name = f"{base_name}{index:02}"
        if not os.path.exists(new_name):
            os.makedirs(new_name)
            return new_name
        index += 1


unique_dir = create_unique_directory('extracted_texts')
input_wb = load_workbook('Input.xlsx')
input_ws = input_wb.active

for row in input_ws.iter_rows(min_row=2,values_only=True):
    url_id,url =row
    title,article_text = extract_data(url)

    with open(os.path.join(unique_dir, f'{url_id}.txt'),'w',encoding='utf-8') as f:
        f.write(f"{title}\n\n{article_text}")


print(f"Data extraction completed.Data's saved at {unique_dir}")


